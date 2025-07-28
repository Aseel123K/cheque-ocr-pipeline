import pandas as pd
from ocr import extract_text_from_image_bytes
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import datetime
from tqdm import tqdm

# === Predict fields for each cheque and export to Excel ===
def predict_and_export(df, models, output_file=None):
    output_rows = []

    print("\nGenerating predictions for Excel output...")
    for i, row in tqdm(df.iterrows(), total=len(df)):
        image_bytes = row.get("image_bytes")
        if not image_bytes:
            continue

        text = extract_text_from_image_bytes(image_bytes)
        if not text.strip():
            continue

        prediction = {"cheque_number": i + 1}
        for field, (model, encoder) in models.items():
            try:
                y_pred = model.predict([text])[0]
                label = encoder.inverse_transform([y_pred])[0]
                prediction[field] = label
            except:
                prediction[field] = "Error"

        output_rows.append(prediction)

    result_df = pd.DataFrame(output_rows)

    # === Auto-generate unique output filename if not provided ===
    if not output_file:
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
        output_file = f"cheque_field_predictions_{timestamp}.xlsx"

    # === Write to Excel with formatting ===
    wb = Workbook()
    ws = wb.active
    ws.title = "Cheque Predictions"

    for r_idx, row in enumerate(dataframe_to_rows(result_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value))
                if cell_len > max_length:
                    max_length = cell_len
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_file)
    print(f"Saved data to {output_file}")
